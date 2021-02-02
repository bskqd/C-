import itertools
from datetime import datetime

from openpyxl import load_workbook
from django.core.management import BaseCommand
from django.db.models import Q, Value
from django.db.models.functions import Concat

from directory.models import Course, NTZ
from sailor.models import Profile
from sailor.document.models import CertificateETI
from communication.models import SailorKeys


STATUS_DOCUMENT_ID = 2


class Command(BaseCommand):
    help = """
        Parse ntz ETIRegistry from xlsx-file
            ./manage.py eti_registry_parser [--input-file filename] \n
            Params:\n
            -i, --input-file : Input xlsx-file
        """

    def add_arguments(self, parser):
        parser.add_argument('-i', '--input-file', dest='input_file', default='./alyans.xlsx', type=str)

    def handle(self, *args, **options):
        wb = load_workbook(options.get('input_file'))
        sheet_name = wb.get_sheet_names()[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        try:
            NTZ_ID = NTZ.objects.get(name__icontains='ТОВ \"Освітній альянс\"').id
        except NTZ.DoesNotExist:
            print('Not found NTZ')
            raise TypeError
        for row in sheet.iter_rows(min_row=2, max_col=7):
            number = row[0].value
            sailor_full_name = row[1].value
            course_find = row[2].value
            date_start = row[4].value
            date_end = row[5].value
            course = Course.objects.filter(name_ukr__icontains=course_find).first()
            qs = Profile.objects.annotate(fullname_ukr=Concat('last_name_ukr', Value(' '), 'first_name_ukr', Value(' '),
                                                              'middle_name_ukr'))
            profile = qs.filter(fullname_ukr__icontains=sailor_full_name)
            if profile:
                for prof in profile:
                    key = SailorKeys.objects.filter(profile=prof.id).first()
                    if not key or not key.sertificate_ntz:
                        continue
                    sailor_cert_exists = CertificateETI.objects.filter(id__in=key.sertificate_ntz,
                                                                       date_start=date_start.date(), ntz_id=NTZ_ID)
                    if not sailor_cert_exists.exists() or sailor_cert_exists.filter(ntz_number=int(number)).exists():
                        continue
                    find_cert = CertificateETI.objects.filter(ntz_number=int(number), date_end=date_end.date(),
                                                              date_start=date_start.date(), ntz_id=NTZ_ID,
                                                              course_training_id=course.id,
                                                              status_document_id=2)
                    if find_cert.exists():
                        find_cert_ids = list(find_cert.values_list('id', flat=True))
                        list_use_certs = list(SailorKeys.objects.filter(sertificate_ntz__overlap=find_cert_ids).
                                              values_list('sertificate_ntz', flat=True))
                        ids_use_cert = list(itertools.chain.from_iterable(list_use_certs))
                        not_use_cert = list(set(find_cert_ids) - (set(find_cert_ids) & set(ids_use_cert)))
                        if not_use_cert:
                            key.sertificate_ntz.append(not_use_cert[0])
                            key.save(update_fields=['sertificate_ntz'])
                            print('used existing certificate')
                            continue
                    cert = CertificateETI.objects.create(ntz_number=number, date_end=date_end.date(),
                                                         date_start=date_start.date(), ntz_id=NTZ_ID,
                                                         course_training_id=course.id,
                                                         status_document_id=STATUS_DOCUMENT_ID,
                                                         )
                    key.sertificate_ntz.append(cert.id)
                    key.save(update_fields=['sertificate_ntz'])
                    print('created new certificate')
            else:
                print(f'{sailor_full_name} not found')


