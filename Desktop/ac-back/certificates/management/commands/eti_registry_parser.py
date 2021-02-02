from openpyxl import load_workbook
from django.core.management import BaseCommand

from directory.models import Course
from certificates.models import ETIRegistry


class Command(BaseCommand):
    help = """
        Parse ntz ETIRegistry from xlsx-file
            ./manage.py eti_registry_parser [--input-file filename] \n
            Params:\n
            -i, --input-file : Input xlsx-file
        """

    def add_arguments(self, parser):
        parser.add_argument('-i', '--input-file', dest='input_file', default='./eti_registry.xlsx', type=str)

    def handle(self, *args, **options):
        course_ids = list(Course.objects.filter(is_disable=False).values('id', 'code_for_parsing'))
        wb = load_workbook(options.get('input_file'))
        sheet_names = wb.get_sheet_names()
        for sheet_name in sheet_names[1:]:
            sheet_data = []
            sheet = wb.get_sheet_by_name(sheet_name)
            ntz_id = sheet['A1'].value
            for row in sheet.iter_rows(min_row=6, max_col=5, max_row=57):
                course_num = row[0].value
                date_start = row[2].value
                if date_start is None:
                    continue
                course_id = \
                    [course['id'] for course in course_ids if course['code_for_parsing'] == float(course_num)][0]
                date_end = row[3].value
                number_with_year = row[4].value
                number_with_year = number_with_year.replace('№', '').strip()
                number_protocol = int(number_with_year.split('/')[0])
                sheet_data.append(ETIRegistry(institution_id=ntz_id, course_id=course_id, date_start=date_start,
                                              date_end=date_end, number_protocol=number_protocol))
            ETIRegistry.objects.bulk_create(sheet_data, ignore_conflicts=True)
