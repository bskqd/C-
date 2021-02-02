import os

from django.conf import settings
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.urls import NoReverseMatch
from django.utils import timezone as tz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.reverse import reverse_lazy

from itcs import celery_app
from notifications.models import UserNotification
from reports.models import ProtocolFiles
import reports.utils


@celery_app.task
def send_report_email(recipient, description, token):
    try:
        url = reverse_lazy('download-report', args=[token])
    except (Exception, NoReverseMatch):
        url = 'http://127.0.0.1:8000/api/v1/sailor/reports/{}/'.format(token)
    message = settings.REPORT_SENDER_MESSAGE.format(url)
    send_mail(
        settings.REPORTS_SENDER_SUBJECT.format(description),
        from_email=settings.REPORTS_SENDER_EMAIL,
        recipient_list=[recipient],
        message='',
        html_message=message,
    )
    # print(settings.REPORTS_SENDER_SUBJECT.format(description))


@celery_app.task(serializer='pickle', time_limit=2000)
def generate_report(list_queryset: list, titles: list, description: str, requester: str, ordering=None, filtering=None):
    col_titles = [t[0] for t in titles]
    attrs = [t[1] for t in titles]
    wb = Workbook()
    ws1 = wb.active

    ws1.title = description
    ws1.append(col_titles)

    for queryset in list_queryset:
        paginator = Paginator(queryset, settings.REPORTS_CHUNK_SIZE)
        for page in range(1, paginator.num_pages + 1):
            for row in paginator.page(page).object_list:
                ws1.append(
                    [getattr(row, attr) for attr in attrs]
                )
    reports.utils.settings_excel_file(ws1)

    if not os.path.exists(settings.MEDIA_ROOT + '/protocols/'):
        os.mkdir(settings.MEDIA_ROOT + '/protocols/')
    description = description.replace(' ', '_')
    file_name = f'{settings.MEDIA_ROOT}/protocols/{tz.now().date().isoformat()}-{description}-' \
                f'{tz.now().timestamp()}.xlsx'
    wb.save(file_name)
    protocol_file = ProtocolFiles(file_path=file_name, user=requester).save()
    # send_report_email.s(requester, description, report.token).apply_async()
    UserNotification.objects.create(recipient=requester,
                                    title='Ваш звіт згенеровано',
                                    text='Для того щоб скачати ваш звіт перейдіть до розд'
                                         'ілу зо звітами та нажміть іконку завантаження',
                                    obj=protocol_file, sailor_id=None)


@celery_app.task(serializer='pickle', time_limit=2000)
def generate_report_back_office(description: str, requester: str, attrs=None):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = description
    titles = [col[0] for col in attrs[0]]
    ws1.append(titles)
    [ws1.append([value[1] for value in attr]) for attr in attrs]

    reports.utils.settings_excel_file(ws1)

    if not os.path.exists(settings.MEDIA_ROOT + '/reports/'):
        os.mkdir(settings.MEDIA_ROOT + '/reports/')
    description = description.replace(' ', '_')
    file_name = f'{settings.MEDIA_ROOT}/reports/{tz.now().date().isoformat()}-{description}-{tz.now().timestamp()}.xlsx'
    wb.save(file_name)
    protocol_file = ProtocolFiles(file_path=file_name, user=requester).save()
    # send_report_email.s(requester, description, report.token).apply_async()
    UserNotification.objects.create(recipient=requester,
                                    title='Ваш звіт згенеровано',
                                    text='Для того щоб скачати ваш звіт перейдіть до розд'
                                         'ілу зо звітами та нажміть іконку завантаження',
                                    obj=protocol_file, sailor_id=None)
