from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def settings_excel_file(ws1):
    ws1.auto_filter.ref = ws1.calculate_dimension()
    for idx in range(ws1.min_column, ws1.max_column + 1):
        col = ws1.column_dimensions[get_column_letter(idx)]
        col.bestFit = True
        col.auto_size = True
        col.width = 26.0
        # col.customWidth = True

    for idx in range(ws1.min_row + 1, ws1.max_row + 1):
        row = ws1.row_dimensions[idx]
        row.alignment = Alignment(vertical='center')
        row.height = 20

    row = ws1.row_dimensions[ws1.min_row]
    row.font = Font(size=12, bold=True, name='Calibri')
    row.fill = PatternFill('solid', start_color='FFCCCC00')
    row.alignment = Alignment(vertical='center', horizontal='center')
    row.height = 25
    return ws1
