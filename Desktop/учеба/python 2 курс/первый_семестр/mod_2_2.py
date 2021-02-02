import openpyxl


def change_xl(data):
    wb = openpyxl.load_workbook(data)

    for sheet_num in range(len(wb.sheetnames)):
        last = []
        draft = ''
        ws = wb.worksheets[sheet_num]
        for row in range(1, 5):  # можна вибрати range такий, який потрібно
            for column in "ABCDE":  # можна вибрати потрібні букви
                value = ws[f"{column}{row}"].value
                if value is not None and value == int(value):
                    last.append(
                        {'column': column, 'row': row}
                    )

        for i, cell in enumerate(last, 1):
            if i == len(last):
                draft += f"{cell['column']}{cell['row']}"
            else:
                draft += f"{cell['column']}{cell['row']}:"
        ws['A7'].value = f"=SUM({draft})"

    wb.save(data)


if __name__ == '__main__':
    change_xl('example.xlsx')
